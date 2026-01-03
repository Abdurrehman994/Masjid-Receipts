
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import Optional, List, Dict
from datetime import datetime
from app.core.database import get_db
from app.models.user import User, UserRole
from app.models.receipt import Receipt, PaymentMode
from app.models.tag import Tag
from app.schemas.receipt import Receipt as ReceiptSchema
from app.api.deps import get_current_user, require_role
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter()


@router.get("/tally")
def get_tally(
    month: Optional[int] = Query(None, ge=1, le=12, description="Filter by month (1-12)"),
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Filter by year"),
    tag_id: Optional[int] = Query(None, description="Filter by tag ID"),
    tag_name: Optional[str] = Query(None, description="Filter by tag name"),
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    ENDPOINT: GET TALLY (CORE FEATURE!)
    WHO CAN USE: Finance Secretary & Auditor ONLY
    """
    
    # Start with base query
    query = db.query(Receipt)
    
    # Build filters
    filters_applied = {}
    
    # Filter by month and year
    if month:
        query = query.filter(extract('month', Receipt.receipt_date) == month)
        filters_applied["month"] = month
    
    if year:
        query = query.filter(extract('year', Receipt.receipt_date) == year)
        filters_applied["year"] = year
    
    # Filter by tag (either by ID or name)
    if tag_id or tag_name:
        if tag_id:
            tag = db.query(Tag).filter(Tag.id == tag_id).first()
            filters_applied["tag_id"] = tag_id
        else:
            tag = db.query(Tag).filter(Tag.name == tag_name).first()
            filters_applied["tag_name"] = tag_name
        
        if tag:
            # Filter receipts that have this tag
            query = query.filter(Receipt.tags.contains(tag))
            filters_applied["tag"] = tag.name
    
    # Get all matching receipts
    receipts = query.all()
    
    # Calculate total amount
    total_amount = sum(receipt.amount for receipt in receipts)
    receipt_count = len(receipts)
    
    # Calculate breakdown by category
    category_breakdown = db.query(
        Receipt.category,
        func.sum(Receipt.amount).label('total'),
        func.count(Receipt.id).label('count')
    ).filter(Receipt.id.in_([r.id for r in receipts] if receipts else [0])).\
    group_by(Receipt.category).all()
    
    by_category = [
        {
            "category": cat,
            "total": float(total),
            "count": count
        }
        for cat, total, count in category_breakdown
    ]
    
    # Calculate breakdown by payment mode
    payment_breakdown = db.query(
        Receipt.payment_mode,
        func.sum(Receipt.amount).label('total'),
        func.count(Receipt.id).label('count')
    ).filter(Receipt.id.in_([r.id for r in receipts] if receipts else [0])).\
    group_by(Receipt.payment_mode).all()
    
    by_payment_mode = [
        {
            "payment_mode": mode.value if mode else "unknown",
            "total": float(total),
            "count": count
        }
        for mode, total, count in payment_breakdown
    ]
    
    return {
        "total_amount": round(total_amount, 2),
        "receipt_count": receipt_count,
        "filters_applied": filters_applied,
        "by_category": sorted(by_category, key=lambda x: x['total'], reverse=True),
        "by_payment_mode": sorted(by_payment_mode, key=lambda x: x['total'], reverse=True)
    }


@router.get("/by-tag/{tag_name}", response_model=List[ReceiptSchema])
def get_receipts_by_tag(
    tag_name: str,
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    ENDPOINT: GET RECEIPTS BY TAG
    
    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Find tag
    tag = db.query(Tag).filter(Tag.name == tag_name).first()
    
    if not tag:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"Tag '{tag_name}' not found")
    
    # Get receipts with this tag
    receipts = db.query(Receipt).filter(Receipt.tags.contains(tag)).all()
    
    return receipts


@router.get("/monthly-breakdown")
def get_monthly_breakdown(
    year: int = Query(..., ge=2000, le=2100, description="Year to analyze"),
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    ENDPOINT: MONTHLY BREAKDOWN
    
    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Get monthly totals
    monthly_data = db.query(
        extract('month', Receipt.receipt_date).label('month'),
        func.sum(Receipt.amount).label('total'),
        func.count(Receipt.id).label('count')
    ).filter(
        extract('year', Receipt.receipt_date) == year
    ).group_by(
        extract('month', Receipt.receipt_date)
    ).all()
    
    # Month names
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    
    # Format results
    months = []
    total_amount = 0
    
    for month_num, total, count in monthly_data:
        month_num = int(month_num)
        total = float(total)
        total_amount += total
        
        months.append({
            "month": month_num,
            "month_name": month_names[month_num - 1],
            "total": round(total, 2),
            "count": count
        })
    
    # Sort by month
    months.sort(key=lambda x: x['month'])
    
    return {
        "year": year,
        "total_amount": round(total_amount, 2),
        "months": months
    }


@router.get("/summary")
def get_summary(
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    ENDPOINT: OVERALL SUMMARY

    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Total stats
    total_receipts = db.query(func.count(Receipt.id)).scalar()
    total_amount = db.query(func.sum(Receipt.amount)).scalar() or 0
    average = (total_amount / total_receipts) if total_receipts > 0 else 0
    
    # Top 5 categories
    top_categories = db.query(
        Receipt.category,
        func.sum(Receipt.amount).label('total'),
        func.count(Receipt.id).label('count')
    ).group_by(Receipt.category)\
     .order_by(func.sum(Receipt.amount).desc())\
     .limit(5).all()
    
    top_cats = [
        {
            "category": cat,
            "total": float(total),
            "count": count
        }
        for cat, total, count in top_categories
    ]
    
    # Receipts by uploader role
    receipts_by_role = {}
    for role in UserRole:
        count = db.query(func.count(Receipt.id))\
                 .join(User, Receipt.uploaded_by == User.id)\
                 .filter(User.role == role)\
                 .scalar()
        if count > 0:
            receipts_by_role[role.value] = count
    
    # Recent 5 receipts
    recent = db.query(Receipt)\
               .order_by(Receipt.created_at.desc())\
               .limit(5).all()
    
    recent_receipts = [
        {
            "id": r.id,
            "amount": r.amount,
            "category": r.category,
            "uploader": r.uploader.username,
            "created_at": r.created_at.isoformat()
        }
        for r in recent
    ]
    
    return {
        "total_receipts": total_receipts,
        "total_amount": round(float(total_amount), 2),
        "average_receipt": round(average, 2),
        "top_categories": top_cats,
        "receipts_by_role": receipts_by_role,
        "recent_receipts": recent_receipts
    }
    
@router.get("/export/receipts")
def export_receipts_to_excel(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    tag_name: Optional[str] = Query(None),
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    BONUS FEATURE 2: EXPORT TO EXCEL
    
    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Build query with filters
    query = db.query(Receipt)
    
    if month:
        query = query.filter(extract('month', Receipt.receipt_date) == month)
    
    if year:
        query = query.filter(extract('year', Receipt.receipt_date) == year)
    
    if tag_name:
        tag = db.query(Tag).filter(Tag.name == tag_name).first()
        if tag:
            query = query.filter(Receipt.tags.contains(tag))
    
    receipts = query.all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Add headers
    headers = ["ID", "Date", "Amount", "Category", "Payment Mode", "Store", "Note", "Uploaded By", "Tags"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    for row, receipt in enumerate(receipts, 2):
        ws.cell(row=row, column=1, value=receipt.id)
        ws.cell(row=row, column=2, value=receipt.receipt_date.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=3, value=receipt.amount)
        ws.cell(row=row, column=4, value=receipt.category)
        ws.cell(row=row, column=5, value=receipt.payment_mode.value if receipt.payment_mode else "")
        ws.cell(row=row, column=6, value=receipt.store_name or "")
        ws.cell(row=row, column=7, value=receipt.note or "")
        ws.cell(row=row, column=8, value=receipt.uploader.full_name)
        ws.cell(row=row, column=9, value=", ".join([tag.name for tag in receipt.tags]))
    
    # Add summary row
    summary_row = len(receipts) + 3
    ws.cell(row=summary_row, column=2, value="TOTAL:")
    ws.cell(row=summary_row, column=2).font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=sum(r.amount for r in receipts))
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    filename = f"masjid_receipts_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/tally")
def export_tally_to_excel(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    tag_name: Optional[str] = Query(None),
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    BONUS FEATURE 2B: EXPORT TALLY TO EXCEL
    
    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Get tally data (reuse the tally endpoint logic)
    query = db.query(Receipt)
    
    if month:
        query = query.filter(extract('month', Receipt.receipt_date) == month)
    if year:
        query = query.filter(extract('year', Receipt.receipt_date) == year)
    if tag_name:
        tag = db.query(Tag).filter(Tag.name == tag_name).first()
        if tag:
            query = query.filter(Receipt.tags.contains(tag))
    
    receipts = query.all()
    total_amount = sum(receipt.amount for receipt in receipts)
    
    # Category breakdown
    category_breakdown = db.query(
        Receipt.category,
        func.sum(Receipt.amount).label('total'),
        func.count(Receipt.id).label('count')
    ).filter(Receipt.id.in_([r.id for r in receipts] if receipts else [0])).\
    group_by(Receipt.category).all()
    
    # Create Excel workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    
    ws1['A1'] = "Masjid Receipts - Tally Report"
    ws1['A1'].font = Font(size=16, bold=True)
    
    ws1['A3'] = "Total Amount:"
    ws1['B3'] = total_amount
    ws1['B3'].font = Font(bold=True, size=14)
    
    ws1['A4'] = "Total Receipts:"
    ws1['B4'] = len(receipts)
    
    ws1['A5'] = "Report Date:"
    ws1['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Sheet 2: By Category
    ws2 = wb.create_sheet("By Category")
    ws2['A1'] = "Category"
    ws2['B1'] = "Total Amount"
    ws2['C1'] = "Count"
    ws2['D1'] = "Percentage"
    
    for col in range(1, 5):
        ws2.cell(row=1, column=col).font = Font(bold=True)
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws2.cell(row=1, column=col).font = Font(color="FFFFFF", bold=True)
    
    for row, (cat, total, count) in enumerate(category_breakdown, 2):
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=float(total))
        ws2.cell(row=row, column=3, value=count)
        percentage = (float(total) / total_amount * 100) if total_amount > 0 else 0
        ws2.cell(row=row, column=4, value=f"{percentage:.1f}%")
    
    # Auto-adjust columns
    for ws in wb:
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"masjid_tally_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    
@router.get("/dashboard/charts")
def get_chart_data(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    current_user: User = Depends(require_role(UserRole.FINANCE_SECRETARY, UserRole.AUDITOR)),
    db: Session = Depends(get_db)
):
    """
    BONUS FEATURE 3: DASHBOARD CHART DATA
    
    WHO CAN USE: Finance Secretary & Auditor
    """
    
    # Build query
    query = db.query(Receipt)
    
    if month:
        query = query.filter(extract('month', Receipt.receipt_date) == month)
    if year:
        query = query.filter(extract('year', Receipt.receipt_date) == year)
    
    receipts = query.all()
    
    if not receipts:
        return {
            "message": "No data available for the selected period",
            "stats": {
                "total_amount": 0,
                "receipt_count": 0
            }
        }
    
    total_amount = sum(r.amount for r in receipts)
    
    # Pie Chart Data: By Category
    category_data = db.query(
        Receipt.category,
        func.sum(Receipt.amount).label('total')
    ).filter(Receipt.id.in_([r.id for r in receipts])).\
    group_by(Receipt.category).all()
    
    category_labels = [cat for cat, _ in category_data]
    category_amounts = [float(total) for _, total in category_data]
    category_colors = ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#FF9F40", "#9966FF", "#FF6384"]
    
    # Pie Chart Data: By Payment Mode
    payment_data = db.query(
        Receipt.payment_mode,
        func.sum(Receipt.amount).label('total')
    ).filter(Receipt.id.in_([r.id for r in receipts])).\
    group_by(Receipt.payment_mode).all()
    
    payment_labels = [mode.value if mode else "unknown" for mode, _ in payment_data]
    payment_amounts = [float(total) for _, total in payment_data]
    payment_colors = ["#4BC0C0", "#FF9F40", "#9966FF", "#FF6384", "#36A2EB"]
    
    # Top Categories
    top_categories = sorted(
        [
            {
                "category": cat,
                "amount": float(total),
                "percentage": round((float(total) / total_amount * 100), 1)
            }
            for cat, total in category_data
        ],
        key=lambda x: x['amount'],
        reverse=True
    )[:5]
    
    # Statistics
    amounts = [r.amount for r in receipts]
    stats = {
        "total_amount": round(total_amount, 2),
        "receipt_count": len(receipts),
        "average_receipt": round(total_amount / len(receipts), 2),
        "largest_expense": round(max(amounts), 2),
        "smallest_expense": round(min(amounts), 2)
    }
    
    # Bar Chart: Monthly trend (if year is specified)
    monthly_trend = None
    if year:
        monthly_data = db.query(
            extract('month', Receipt.receipt_date).label('month'),
            func.sum(Receipt.amount).label('total')
        ).filter(
            extract('year', Receipt.receipt_date) == year
        ).group_by(
            extract('month', Receipt.receipt_date)
        ).all()
        
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        monthly_trend = {
            "labels": [month_names[int(m)-1] for m, _ in monthly_data],
            "data": [float(total) for _, total in monthly_data]
        }
    
    return {
        "pie_chart_category": {
            "labels": category_labels,
            "data": category_amounts,
            "colors": category_colors[:len(category_labels)]
        },
        "pie_chart_payment": {
            "labels": payment_labels,
            "data": payment_amounts,
            "colors": payment_colors[:len(payment_labels)]
        },
        "bar_chart_monthly": monthly_trend,
        "top_categories": top_categories,
        "stats": stats
    }